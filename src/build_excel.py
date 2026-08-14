"""
Builds the Excel underwriting model from the pipeline's output.

Usage:
    python src/build_excel.py

Produces output/credit_underwriting_model.xlsx with:
    Summary          one row per company: metrics, grade, watchlist status
    Company Detail   full write-up for a strong, a borderline and a weak name
    Stress Scenarios company x scenario leverage and coverage matrices
    Assumptions      every threshold and assumption, with its rationale
    Data Quality     which concept fed each metric, and what was missing

The covenant-headroom and breach columns on the Summary tab are written as
live Excel formulas pointing at the Assumptions tab, not as pasted values, so
a reviewer can change the assumed covenant in one cell and watch the model
respond -- which is the point of an underwriting model rather than a report.
"""
import sys
from pathlib import Path

import pandas as pd
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import (
    ALTMAN_ZPP_DISTRESS,
    ALTMAN_ZPP_SAFE,
    CASH_TAX_RATE,
    DEBT_EBITDA_COVENANT,
    EBITDA_TO_FCF_PASSTHROUGH,
    EW_EBITDA_DECLINE_PCT,
    EW_LEVERAGE_RISE_TURNS,
    EW_MARGIN_DROP_BPS,
    FLOATING_RATE_DEBT_PCT,
    INVENTORY_STRESS_PCT,
    MIN_CURRENT_RATIO,
    MIN_FCF_TO_DEBT,
    MIN_INTEREST_COVERAGE,
    MIN_SCORECARD_COVERAGE,
    REFI_CONCENTRATION_PCT,
    SCORECARD,
)

OUTPUT_DIR = Path(__file__).parent.parent / "output"
WORKBOOK_PATH = OUTPUT_DIR / "credit_underwriting_model.xlsx"

NAVY = "1F3864"
GREY = "F2F2F2"
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
AMBER_FILL = PatternFill("solid", fgColor="FFEB9C")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
THIN = Side(style="thin", color="BFBFBF")

MONEY = '#,##0,,"m";[Red](#,##0,,"m")'
MULTIPLE = '0.00"x";[Red](0.00"x")'
PERCENT = "0.0%;[Red](0.0%)"
NUMBER = "0.00;[Red](0.00)"


def style_header(worksheet, row: int, last_column: int, fill: str = NAVY) -> None:
    for col in range(1, last_column + 1):
        cell = worksheet.cell(row=row, column=col)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN)


def title(worksheet, text: str, subtitle: str = "") -> None:
    worksheet["A1"] = text
    worksheet["A1"].font = Font(bold=True, size=14, color=NAVY)
    if subtitle:
        worksheet["A2"] = subtitle
        worksheet["A2"].font = Font(italic=True, size=9, color="595959")


def autosize(worksheet, minimum: int = 10, maximum: int = 46) -> None:
    for column_cells in worksheet.columns:
        letter = get_column_letter(column_cells[0].column)
        longest = max((len(str(c.value)) for c in column_cells if c.value is not None), default=0)
        worksheet.column_dimensions[letter].width = max(minimum, min(maximum, longest + 2))


def write_table(worksheet, frame: pd.DataFrame, start_row: int, headers: list[str] | None = None,
                number_formats: dict[str, str] | None = None) -> int:
    """Write a dataframe as a formatted block; returns the last row used."""
    headers = headers or list(frame.columns)
    number_formats = number_formats or {}
    for offset, header in enumerate(headers, start=1):
        worksheet.cell(row=start_row, column=offset, value=header)
    style_header(worksheet, start_row, len(headers))

    for row_offset, (_, record) in enumerate(frame.iterrows(), start=start_row + 1):
        for col_offset, column_name in enumerate(frame.columns, start=1):
            value = record[column_name]
            if pd.isna(value):
                value = None
            elif hasattr(value, "item"):
                value = value.item()
            cell = worksheet.cell(row=row_offset, column=col_offset, value=value)
            cell.border = Border(bottom=THIN)
            if column_name in number_formats:
                cell.number_format = number_formats[column_name]
            if row_offset % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=GREY)
    return start_row + len(frame)


# ---------------------------------------------------------------------------
# Tab 1: Summary / watchlist
# ---------------------------------------------------------------------------
SUMMARY_FIELDS = [
    ("ticker", "Ticker", None),
    ("company_name", "Company", None),
    ("period_end_date", "Period end", "yyyy-mm-dd"),
    ("credit_grade", "Internal grade", None),
    ("pd_proxy_band", "PD proxy band", None),
    ("watchlist_severity", "Watchlist", None),
    ("revenue_ttm", "Revenue TTM", MONEY),
    ("ebitda_ttm", "EBITDA TTM", MONEY),
    ("operating_margin", "Op margin", PERCENT),
    ("total_debt", "Total debt", MONEY),
    ("net_debt", "Net debt", MONEY),
    ("debt_to_ebitda", "Debt/EBITDA", MULTIPLE),
    ("net_debt_to_ebitda", "Net debt/EBITDA", MULTIPLE),
    ("interest_coverage", "EBITDA/interest", MULTIPLE),
    ("fcf_ttm", "FCF TTM", MONEY),
    ("fcf_to_debt", "FCF/debt", PERCENT),
    ("current_ratio", "Current ratio", MULTIPLE),
    ("quick_ratio", "Quick ratio", MULTIPLE),
    ("working_capital", "Working capital", MONEY),
    ("altman_z_double_prime", 'Altman Z"', NUMBER),
    ("altman_zone", "Z\" zone", None),
    ("pct_debt_due_within_1y", "Debt due <1y", PERCENT),
    ("scorecard_coverage", "Data coverage", PERCENT),
]


def build_summary_tab(workbook, summary: pd.DataFrame, watchlist: pd.DataFrame) -> None:
    worksheet = workbook.create_sheet("Summary")
    title(worksheet, "Corporate Credit Portfolio - Summary & Watchlist",
          "All figures from SEC EDGAR XBRL filings. Covenant thresholds are stated "
          "assumptions (see Assumptions tab), not real negotiated terms.")

    merged = summary.merge(
        watchlist[["ticker", "watchlist_severity", "all_triggers"]], on="ticker", how="left")
    merged = merged.sort_values(["debt_to_ebitda"], ascending=False)

    frame = pd.DataFrame({label: merged[field] if field in merged.columns else None
                          for field, label, _ in SUMMARY_FIELDS})
    formats = {label: fmt for _, label, fmt in SUMMARY_FIELDS if fmt}
    last_row = write_table(worksheet, frame, start_row=4, number_formats=formats)

    columns = {label: idx for idx, (_, label, _) in enumerate(SUMMARY_FIELDS, start=1)}

    # Live formulas: headroom is computed by Excel off the Assumptions tab, so
    # changing the assumed covenant re-rates the whole portfolio in place.
    headroom_col = len(SUMMARY_FIELDS) + 1
    breach_col = headroom_col + 1
    worksheet.cell(row=4, column=headroom_col, value="Covenant headroom (turns)")
    worksheet.cell(row=4, column=breach_col, value="Covenant status")
    style_header(worksheet, 4, breach_col)

    leverage_letter = get_column_letter(columns["Debt/EBITDA"])
    for row in range(5, last_row + 1):
        headroom = worksheet.cell(row=row, column=headroom_col)
        headroom.value = (f"=IF(ISBLANK({leverage_letter}{row}),\"n/a\","
                          f"Assumptions!$C$5-{leverage_letter}{row})")
        headroom.number_format = NUMBER
        status = worksheet.cell(row=row, column=breach_col)
        status.value = (f"=IF(ISBLANK({leverage_letter}{row}),\"no data\","
                        f"IF({leverage_letter}{row}>Assumptions!$C$5,\"BREACH\",\"Compliant\"))")
        status.alignment = Alignment(horizontal="center")

    # Traffic-light rules mirror the credit policy thresholds exactly.
    def rule_range(label: str) -> str:
        letter = get_column_letter(columns[label])
        return f"{letter}5:{letter}{last_row}"

    worksheet.conditional_formatting.add(
        rule_range("Debt/EBITDA"),
        CellIsRule(operator="greaterThan", formula=[str(DEBT_EBITDA_COVENANT)], fill=RED_FILL))
    worksheet.conditional_formatting.add(
        rule_range("Debt/EBITDA"),
        CellIsRule(operator="between", formula=["2", str(DEBT_EBITDA_COVENANT)], fill=AMBER_FILL))
    worksheet.conditional_formatting.add(
        rule_range("Debt/EBITDA"),
        CellIsRule(operator="lessThanOrEqual", formula=["2"], fill=GREEN_FILL))
    worksheet.conditional_formatting.add(
        rule_range("EBITDA/interest"),
        CellIsRule(operator="lessThan", formula=[str(MIN_INTEREST_COVERAGE)], fill=RED_FILL))
    worksheet.conditional_formatting.add(
        rule_range("Current ratio"),
        CellIsRule(operator="lessThan", formula=[str(MIN_CURRENT_RATIO)], fill=RED_FILL))
    worksheet.conditional_formatting.add(
        rule_range("FCF/debt"),
        CellIsRule(operator="lessThan", formula=[str(MIN_FCF_TO_DEBT)], fill=AMBER_FILL))
    worksheet.conditional_formatting.add(
        rule_range('Altman Z"'),
        CellIsRule(operator="lessThan", formula=[str(ALTMAN_ZPP_DISTRESS)], fill=RED_FILL))
    worksheet.conditional_formatting.add(
        rule_range('Altman Z"'),
        CellIsRule(operator="greaterThan", formula=[str(ALTMAN_ZPP_SAFE)], fill=GREEN_FILL))
    worksheet.conditional_formatting.add(
        rule_range("Watchlist"),
        CellIsRule(operator="equal", formula=['"High"'], fill=RED_FILL))
    worksheet.conditional_formatting.add(
        rule_range("Watchlist"),
        CellIsRule(operator="equal", formula=['"Medium"'], fill=AMBER_FILL))
    worksheet.conditional_formatting.add(
        rule_range("Data coverage"),
        ColorScaleRule(start_type="num", start_value=0.5, start_color="FFC7CE",
                       end_type="num", end_value=1.0, end_color="C6EFCE"))

    worksheet.freeze_panes = "C5"
    autosize(worksheet)

    note_row = last_row + 2
    worksheet.cell(row=note_row, column=1,
                   value="Watchlist triggers by name (Level = breach today, "
                         "Trend = deteriorating, Stress = breaches under scenario):")
    worksheet.cell(row=note_row, column=1).font = Font(bold=True, size=10)
    for offset, (_, record) in enumerate(
            merged[merged["all_triggers"].notna() & (merged["all_triggers"] != "")].iterrows(), start=1):
        cell = worksheet.cell(row=note_row + offset, column=1,
                              value=f"{record['ticker']}: {record['all_triggers']}")
        cell.font = Font(size=9)
        cell.alignment = Alignment(wrap_text=False)


# ---------------------------------------------------------------------------
# Tab 2: Company detail
# ---------------------------------------------------------------------------
def pick_detail_companies(summary: pd.DataFrame, watchlist: pd.DataFrame) -> list[tuple[str, str]]:
    """One clearly-strong, one borderline and one clearly-weak name, chosen
    from the data rather than hardcoded, so the tab stays right if the
    coverage universe changes."""
    ranked = summary.dropna(subset=["scorecard_score"]).sort_values("scorecard_score")
    if ranked.empty:
        return []
    flagged = set(watchlist.loc[watchlist["on_watchlist"], "ticker"])
    weakest = ranked.iloc[-1]["ticker"]
    strongest = ranked.iloc[0]["ticker"]
    middle = [t for t in ranked["ticker"] if t not in (weakest, strongest) and t in flagged]
    borderline = middle[len(middle) // 2] if middle else ranked.iloc[len(ranked) // 2]["ticker"]
    return [(strongest, "Strongest credit in the portfolio"),
            (borderline, "Borderline / on watchlist"),
            (weakest, "Weakest credit in the portfolio")]


def build_detail_tab(workbook, summary: pd.DataFrame, watchlist: pd.DataFrame,
                     trends: pd.DataFrame, stress: pd.DataFrame) -> None:
    worksheet = workbook.create_sheet("Company Detail")
    title(worksheet, "Company Underwriting Detail",
          "One page per name: current position, eight-quarter trend, stress results "
          "and the triggers driving the recommendation.")

    row = 4
    for ticker, rationale in pick_detail_companies(summary, watchlist):
        company = summary[summary["ticker"] == ticker].iloc[0]
        flags = watchlist[watchlist["ticker"] == ticker].iloc[0]

        worksheet.cell(row=row, column=1,
                       value=f"{ticker} - {company.get('company_name', '')}")
        worksheet.cell(row=row, column=1).font = Font(bold=True, size=12, color=NAVY)
        worksheet.cell(row=row, column=4, value=rationale).font = Font(italic=True, size=9)
        row += 1

        recommendation = {
            "High": "REJECT / REMEDIATE - outside credit policy today",
            "Medium": "APPROVE WITH CONDITIONS - monitor quarterly against triggers below",
            "Low": "APPROVE WITH CONDITIONS - minor deterioration noted",
            "Clear": "APPROVE - within all stated credit policy thresholds",
        }[flags.get("watchlist_severity") or "Clear"]
        worksheet.cell(row=row, column=1, value="Recommendation:").font = Font(bold=True, size=10)
        worksheet.cell(row=row, column=2, value=recommendation)
        row += 1
        worksheet.cell(row=row, column=1, value="Internal grade:").font = Font(bold=True, size=10)
        worksheet.cell(row=row, column=2,
                       value=f"{company['credit_grade']} (PD proxy band: {company['pd_proxy_band']}; "
                             f"scorecard {company['scorecard_score']:.2f} on a 0=best / 2=worst scale)")
        row += 2

        key_metrics = pd.DataFrame({
            "Metric": ["Revenue TTM", "EBITDA TTM", "Operating margin", "Total debt",
                       "Debt/EBITDA", "Net debt/EBITDA", "EBITDA/interest", "FCF/debt",
                       "Current ratio", "Quick ratio", 'Altman Z"'],
            "Actual": [company.get(k) for k in
                       ["revenue_ttm", "ebitda_ttm", "operating_margin", "total_debt",
                        "debt_to_ebitda", "net_debt_to_ebitda", "interest_coverage",
                        "fcf_to_debt", "current_ratio", "quick_ratio", "altman_z_double_prime"]],
            "Policy threshold": ["-", "-", "-", "-", f"<= {DEBT_EBITDA_COVENANT}x", "-",
                                 f">= {MIN_INTEREST_COVERAGE}x", f">= {MIN_FCF_TO_DEBT:.0%}",
                                 f">= {MIN_CURRENT_RATIO}x", "-", f">= {ALTMAN_ZPP_DISTRESS}"],
        })
        row = write_table(worksheet, key_metrics, row) + 2

        company_trend = trends[trends["ticker"] == ticker]
        if not company_trend.empty:
            pivot = company_trend.pivot_table(index="metric", columns="period_end_date",
                                              values="value", aggfunc="last")
            pivot = pivot.reindex(columns=sorted(pivot.columns)[-8:])
            pivot.columns = [str(c)[:10] for c in pivot.columns]
            worksheet.cell(row=row, column=1, value="Eight-quarter trend").font = Font(bold=True, size=10)
            row += 1
            row = write_table(worksheet, pivot.reset_index().rename(columns={"metric": "Metric"}), row) + 2

        company_stress = stress[stress["ticker"] == ticker][
            ["scenario_label", "stressed_ebitda", "stressed_debt_to_ebitda",
             "stressed_interest_coverage", "stressed_fcf_to_debt", "stressed_quick_ratio"]]
        company_stress.columns = ["Scenario", "Stressed EBITDA", "Debt/EBITDA",
                                  "EBITDA/interest", "FCF/debt", "Quick ratio"]
        worksheet.cell(row=row, column=1, value="Stress scenarios").font = Font(bold=True, size=10)
        row += 1
        row = write_table(worksheet, company_stress, row, number_formats={
            "Stressed EBITDA": MONEY, "Debt/EBITDA": MULTIPLE,
            "EBITDA/interest": MULTIPLE, "FCF/debt": PERCENT, "Quick ratio": MULTIPLE}) + 2

        worksheet.cell(row=row, column=1, value="Triggers").font = Font(bold=True, size=10)
        row += 1
        triggers = flags["all_triggers"] or "None - within all stated thresholds."
        for reason in str(triggers).split(" | "):
            worksheet.cell(row=row, column=1, value=f"- {reason}").font = Font(size=9)
            row += 1
        row += 2

    autosize(worksheet, maximum=34)


# ---------------------------------------------------------------------------
# Tab 3: Stress scenarios
# ---------------------------------------------------------------------------
def build_stress_tab(workbook, stress: pd.DataFrame) -> None:
    worksheet = workbook.create_sheet("Stress Scenarios")
    title(worksheet, "Stress Scenario Matrix",
          f"Assumed covenant {DEBT_EBITDA_COVENANT}x Debt/EBITDA and {MIN_INTEREST_COVERAGE}x "
          f"minimum EBITDA/interest. Red = breach.")

    row = 4
    for label, value_column, fmt, rule in [
        ("Stressed Debt/EBITDA", "stressed_debt_to_ebitda", MULTIPLE,
         CellIsRule(operator="greaterThan", formula=[str(DEBT_EBITDA_COVENANT)], fill=RED_FILL)),
        ("Stressed EBITDA/interest coverage", "stressed_interest_coverage", MULTIPLE,
         CellIsRule(operator="lessThan", formula=[str(MIN_INTEREST_COVERAGE)], fill=RED_FILL)),
        ("Stressed FCF/debt", "stressed_fcf_to_debt", PERCENT,
         CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL)),
        ("Stressed quick ratio", "stressed_quick_ratio", MULTIPLE,
         CellIsRule(operator="lessThan", formula=["0.4"], fill=AMBER_FILL)),
    ]:
        worksheet.cell(row=row, column=1, value=label).font = Font(bold=True, size=11, color=NAVY)
        row += 1
        matrix = stress.pivot_table(index="ticker", columns="scenario_label",
                                    values=value_column, aggfunc="last")
        ordered = [c for c in stress["scenario_label"].unique() if c in matrix.columns]
        matrix = matrix[ordered].reset_index().rename(columns={"ticker": "Ticker"})
        start = row
        row = write_table(worksheet, matrix, row,
                          number_formats={c: fmt for c in matrix.columns if c != "Ticker"})
        span = f"B{start + 1}:{get_column_letter(len(matrix.columns))}{row}"
        worksheet.conditional_formatting.add(span, rule)
        row += 2

    worksheet.freeze_panes = "B5"
    autosize(worksheet, maximum=26)


# ---------------------------------------------------------------------------
# Tab 4: Assumptions
# ---------------------------------------------------------------------------
def build_assumptions_tab(workbook) -> None:
    worksheet = workbook.create_sheet("Assumptions")
    title(worksheet, "Assumptions & Model Limitations",
          "Nothing in this model is a real negotiated covenant. Every threshold below "
          "is an explicit assumption and is the only place it is defined.")

    rows = [
        ("Credit policy", "Max Debt/EBITDA (assumed covenant)", DEBT_EBITDA_COVENANT,
         "Illustrative of total-leverage maintenance tests in public retail term loans. "
         "Real agreements are private; this is not one of them."),
        ("Credit policy", "Min EBITDA/interest coverage", MIN_INTEREST_COVERAGE,
         "Typical minimum coverage test."),
        ("Credit policy", "Min current ratio", MIN_CURRENT_RATIO,
         "Working-capital liquidity floor."),
        ("Credit policy", "Min FCF/debt", MIN_FCF_TO_DEBT,
         "Below this, the company cannot meaningfully deleverage from cash flow."),
        ("Credit policy", "Refinancing concentration trigger", REFI_CONCENTRATION_PCT,
         "Share of total debt maturing within 12 months that warrants a flag."),
        ("Altman", 'Z" distress threshold', ALTMAN_ZPP_DISTRESS,
         "Z\" (non-manufacturer revision) is used, not the 1968 manufacturing Z."),
        ("Altman", 'Z" safe threshold', ALTMAN_ZPP_SAFE, "Above this = safe zone."),
        ("Stress", "Floating-rate share of debt", FLOATING_RATE_DEBT_PCT,
         "Filings rarely give a clean fixed/floating split; assumed, and only this "
         "share reprices in the rate scenario."),
        ("Stress", "Inventory build", INVENTORY_STRESS_PCT,
         "Assumed cash-funded, so current assets are unchanged and the quick ratio "
         "absorbs the hit."),
        ("Stress", "EBITDA-to-FCF passthrough", EBITDA_TO_FCF_PASSTHROUGH,
         "A sales decline releases working capital, cushioning cash flow, so less than "
         "100% of an EBITDA fall reaches FCF in year one."),
        ("Stress", "Cash tax rate", CASH_TAX_RATE,
         "Converts the pre-tax earnings shock into an after-tax cash-flow shock."),
        ("Early warning", "Leverage rise (turns YoY)", EW_LEVERAGE_RISE_TURNS,
         "Trend trigger: still compliant, but moving the wrong way."),
        ("Early warning", "Margin drop (bps YoY)", EW_MARGIN_DROP_BPS, "Trend trigger."),
        ("Early warning", "EBITDA decline (YoY)", EW_EBITDA_DECLINE_PCT, "Trend trigger."),
        ("Scorecard", "Min data coverage for a grade", MIN_SCORECARD_COVERAGE,
         "Below this share of scorecard weight the grade is suppressed as NR, so a "
         "thin data set cannot produce a flattering grade."),
    ]
    frame = pd.DataFrame(rows, columns=["Category", "Assumption", "Value", "Rationale"])
    last_row = write_table(worksheet, frame, 4)
    for row in range(5, last_row + 1):
        worksheet.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical="top")

    row = last_row + 2
    worksheet.cell(row=row, column=1, value="Scorecard weights").font = Font(bold=True, size=11, color=NAVY)
    row += 1
    weights = pd.DataFrame(
        [(metric, direction.replace("_", " "), f"{good} / {weak}", weight)
         for metric, direction, (good, weak), weight in SCORECARD],
        columns=["Metric", "Direction", "Good / weak band", "Weight"])
    row = write_table(worksheet, weights, row, number_formats={"Weight": "0%"}) + 2

    limitations = [
        "Altman Z\" was calibrated on non-manufacturers but not on modern e-commerce "
        "or omni-channel retail; treat the zone as a ranking, not a verdict.",
        "The PD proxy is a rules-based scorecard. It was NOT fit to historical default "
        "data, so its Low/Medium/High bands are relative risk ranking, not probabilities.",
        "Covenant thresholds are assumptions. Real agreements often use net leverage, "
        "adjusted EBITDA with addbacks, and step-downs, all of which would change headroom.",
        "Operating lease liabilities are excluded from total debt. Rating agencies often "
        "capitalise them, which would raise leverage materially for retailers.",
        "EBITDA is unadjusted: no addbacks for impairments, restructuring or stock "
        "compensation, so it is more conservative than a credit agreement's definition.",
        "Where a company discloses only net interest quarterly, coverage is computed on "
        "a net basis and flagged in the interest_basis column.",
        "The model uses the latest filed quarter, which lags the market by up to a "
        "quarter; it cannot see events since the last filing.",
    ]
    worksheet.cell(row=row, column=1, value="Known limitations").font = Font(bold=True, size=11, color=NAVY)
    row += 1
    for item in limitations:
        cell = worksheet.cell(row=row, column=1, value=f"- {item}")
        cell.font = Font(size=9)
        row += 1

    autosize(worksheet, maximum=30)
    worksheet.column_dimensions["D"].width = 70
    worksheet.column_dimensions["A"].width = 60


# ---------------------------------------------------------------------------
# Tab 5: Data quality
# ---------------------------------------------------------------------------
def build_data_quality_tab(workbook, summary: pd.DataFrame) -> None:
    import sqlite3
    worksheet = workbook.create_sheet("Data Quality")
    title(worksheet, "Extraction Audit",
          "Which XBRL concept fed each metric, how much history it covers, and how much "
          "had to be reconstructed from year-to-date disclosures.")

    db_path = Path(__file__).parent.parent / "db" / "credit_risk.db"
    with sqlite3.connect(db_path) as conn:
        quality = pd.read_sql(
            "SELECT ticker AS Ticker, metric AS Metric, tags_used AS 'Concept(s) used', "
            "quarters_found AS Quarters, first_period AS 'First period', "
            "last_period AS 'Last period', derived_share AS 'Share de-cumulated' "
            "FROM data_quality ORDER BY ticker, metric", conn)

    provenance = summary[["ticker", "debt_source", "ebit_source", "interest_basis",
                          "scorecard_coverage", "balance_sheet_as_of"]].copy()
    provenance.columns = ["Ticker", "Debt source", "EBIT source", "Interest basis",
                          "Scorecard coverage", "Balance sheet as of"]
    worksheet.cell(row=4, column=1, value="Input provenance by company").font = Font(bold=True, size=11, color=NAVY)
    row = write_table(worksheet, provenance, 5,
                      number_formats={"Scorecard coverage": "0%"}) + 2

    worksheet.cell(row=row, column=1, value="Concept-level extraction log").font = Font(bold=True, size=11, color=NAVY)
    write_table(worksheet, quality, row + 1, number_formats={"Share de-cumulated": "0%"})
    autosize(worksheet, maximum=44)


def main() -> int:
    from openpyxl import Workbook

    required = ["metrics_summary.csv", "watchlist.csv", "stress_test_results.csv", "trends.csv"]
    missing = [f for f in required if not (OUTPUT_DIR / f).exists()]
    if missing:
        print(f"Missing {', '.join(missing)}. Run the pipeline first.")
        return 1

    summary = pd.read_csv(OUTPUT_DIR / "metrics_summary.csv", parse_dates=["period_end_date"])
    watchlist = pd.read_csv(OUTPUT_DIR / "watchlist.csv")
    stress = pd.read_csv(OUTPUT_DIR / "stress_test_results.csv")
    trends = pd.read_csv(OUTPUT_DIR / "trends.csv", parse_dates=["period_end_date"])

    workbook = Workbook()
    workbook.remove(workbook.active)
    build_summary_tab(workbook, summary, watchlist)
    build_detail_tab(workbook, summary, watchlist, trends, stress)
    build_stress_tab(workbook, stress)
    build_assumptions_tab(workbook)
    build_data_quality_tab(workbook, summary)
    workbook.save(WORKBOOK_PATH)

    print(f"Built {WORKBOOK_PATH}")
    print(f"  Tabs: {', '.join(workbook.sheetnames)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
